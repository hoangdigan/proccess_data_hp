import openpyxl as xl
import os
import glob
from openpyxl import Workbook
from openpyxl import load_workbook

# opening the template vietstock excel file 
filename ="ratio.xlsx"
wb1 = xl.load_workbook(filename)
ws3= wb1['ratio']
ws1 = wb1['calculate']

c=3
for name in glob.glob('D:/ImportDataFinance-UDEMY/process_data/bkhdt/data_vietstock/data_proccess/*.xlsx'): 
   
    wb2 = xl.load_workbook(name, data_only=True)
    ws2= wb2['FSA']
    
    r= 3  
    for i in range(1,15):       
        ws1.cell(2,c).value = ws2.cell(1,2).value
       
        r= r+1
        if i ==1 :
            ws1.cell(r-1,2).value = ws3.cell(i, 1).value
            for t in range(3,8):
                if ws2.cell(12, t).value != None and ws2.cell(7, t).value != None:
                    ws1.cell(r,2).value = ws2.cell(5, t).value 
                    if ws2.cell(7, t).value != 0:
                        ws1.cell(r,c).value = ws2.cell(12, t).value / ws2.cell(7, t).value
                r+=1
        else:
            ws1.cell(r-1,2).value = ws3.cell(i, 1).value
            for t in range(14,19):
                ws1.cell(r,2).value = ws2.cell(5, t).value
                ws1.cell(r,c).value = ws2.cell(ws3.cell(i,2).value, t).value
                r+=1
    c+=1
    wb2.close()

filename_output = "ratio" + '_proccess.xlsx'    
wb1.save(str(filename_output))
    