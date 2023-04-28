import openpyxl as xl
import os
import glob
from openpyxl import Workbook
from openpyxl import load_workbook

# opening the template vietstock excel file 
filename ="vietstock_template.xlsx"
wb1 = xl.load_workbook(filename)
ws1= wb1['MCK']

i=1
for name in glob.glob('D:/ImportDataFinance-UDEMY/process_data/bkhdt/data_vietstock/*.xlsx'): 
   
    ws1.cell(i,1).value = name[78:81]
    ws1.cell(i,2).value = name[100:104]
    i=i+1

filename_output = "D:/ImportDataFinance-UDEMY/process_data/bkhdt/data_vietstock/data_proccess/"+"MACK" + '_proccessed.xlsx'    
wb1.save(str(filename_output))
    