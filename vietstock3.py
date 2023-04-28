# check period data from folder
import openpyxl as xl
import os
from openpyxl import Workbook
from openpyxl import load_workbook

filename = "vietstock_proccess_mapping.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1['MST']
ws3 = wb1['BCTC']  

#run each company
for i in range (2, 3):
    # opening the destination excel file 
    filename1 ="template.xlsx"
    wb2 = xl.load_workbook(filename1)
    ws2 = wb2['BS.data']
    ws4 = wb2['PL.data']
    ws5 = wb2['CF.data']
    ws6 = wb2['FSA']
    # run each company
    company = ws1.cell(row = i, column = 4).value
    company_name = ws1.cell(row = i, column = 2).value
   
    # get year list
    year_list = ["2018", "2019", "2020", "2021", "2022"]

    # for j in range (2, 7873):   
    #     if ws3.cell(j, 8).value == company:
    #         if ws3.cell(j, 2).value not in year_list:
    #             year_list.append(ws3.cell(j, 2).value)

    # run each year for each company
    col = 5  
    for y in year_list:
        ws6.cell(1,2).value = company_name
        for t in range (2, 236):
            for k in range(4, 120): 
            # Update balance sheet
                if ws3.cell(t, 3).value == company and ws3.cell(t, 1).value == 1:  
                    
                    ws2.cell(row = 2, column = col).value = y            
                        
                    if str(ws2.cell(k,2).value) == str(ws3.cell(t,2).value):                    
                        # writing the read value to destination excel file
                        ws2.cell(row = k, column = col).value = ws3.cell(t,col).value

            # Update PL
            for k in range(3, 26): 
                if ws3.cell(t, 3).value == company and ws3.cell(t, 1).value == 2:  
                    ws4.cell(row = 2, column = col).value = y            
                        
                    if str(ws4.cell(k,2).value) == str(ws3.cell(t,2).value):                    
                        # writing the read value to destination excel file
                        ws4.cell(row = k, column = col).value = ws3.cell(t,col).value

            # Update LCTTGT
            for k in range(4, 44): 
                if ws3.cell(t, 3).value == company and ws3.cell(t, 1).value == 5:  
                    ws5.cell(row = 2, column = col).value = y            
                        
                    if str(ws5.cell(k,2).value) == str(ws3.cell(t,2).value): 
                        print("check 20", ws3.cell(t,2).value)                   
                        # writing the read value to destination excel file
                        ws5.cell(row = k, column = col).value = ws3.cell(t,col).value
            
            # Update LCTTTT
            for k in range(48, 85): 
                if ws3.cell(t, 3).value == company and ws3.cell(t, 1).value == 4:  
                    ws5.cell(row = 2, column = col).value = y            
                        
                    if str(ws5.cell(k,2).value) == str(ws3.cell(t,2).value):                    
                        # writing the read value to destination excel file
                        ws5.cell(row = k, column = col).value = ws3.cell(t,col).value
        col +=1     
   
        # saving the destination excel file  
        folder_export = 'D:/ImportDataFinance-UDEMY/process_data/bkhdt/data_vietstock/data_proccess/'
        # D:\ImportDataFinance-UDEMY\process_data\bkhdt\data_vietstock\data_proccess
        
        filename_output = folder_export+ company_name + '_proccess.xlsx'    
        wb2.save(str(filename_output))        
        wb2.close()