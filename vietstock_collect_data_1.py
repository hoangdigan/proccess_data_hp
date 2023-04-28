import openpyxl as xl
import os
import glob
from openpyxl import Workbook
from openpyxl import load_workbook

# opening the template vietstock excel file 
filename ="vietstock_template.xlsx"
wb1 = xl.load_workbook(filename)
ws1= wb1['MST']
ws2 = wb1['BCTC']

for row in ws1['A1:D100']:
  for cell in row:
    cell.value = None

# get stock list
stock_list =[]
for name in glob.glob('D:/ImportDataFinance-UDEMY/process_data/bkhdt/data_vietstock/*.xlsx'): 
    mack = name[78:81]
    if mack not in stock_list:
        stock_list.append(mack)

# update to list to run
t=2
for item in stock_list:
    ws1.cell(t,2).value = item
    ws1.cell(t,4).value = item
    t=t+1

i=1
for name in glob.glob('D:/ImportDataFinance-UDEMY/process_data/bkhdt/data_vietstock/*.xlsx'):  
    # CDKT
    mack = name[78:81]
    if name[100:104] == "CDKT":
        wb2 = xl.load_workbook(name)
        ws3 = wb2['CDKT']
        for j in range(1,7):
            ws2.cell(i,j+3).value = ws3.cell(6,j).value

        i=i+1    
        for k in range(13, 136):
            for j in range(1,7):
                ws2.cell(i,1).value =  1
                ws2.cell(i,3).value =  mack
                ws2.cell(i,j+3).value =  ws3.cell(k,j).value
            i+=1
        wb2.close()

    if name[100:104] == "KQKD":
        wb2 = xl.load_workbook(name)
        ws3 = wb2['KQKD']
        for j in range(1,7):
            ws2.cell(i,j+3).value = ws3.cell(6,j).value

        i=i+1    
        for k in range(13, 37):
            for j in range(1,7):
                ws2.cell(i,1).value =  2
                ws2.cell(i,3).value =  mack
                ws2.cell(i,j+3).value =  ws3.cell(k,j).value
            i+=1
        wb2.close()

    if name[100:104] == "LCTT":
         wb2 = xl.load_workbook(name)
         for ws in wb2:
            if ws.cell(6,1).value == "LƯU CHUYỂN TIỀN TỆ TRỰC TIẾP":
                for j in range(1,7):
                    ws2.cell(i,j+3).value = ws.cell(6,j).value

                i=i+1    
                for k in range(13, 47):
                    for j in range(1,7):
                        ws2.cell(i,1).value =  4
                        ws2.cell(i,3).value =  mack
                        ws2.cell(i,j+3).value =  ws.cell(k,j).value
                    i+=1

            if ws.cell(6,1).value == "LƯU CHUYỂN TIỀN TỆ GIÁN TIẾP":
                for j in range(1,7):
                    ws2.cell(i,j+3).value = ws.cell(6,j).value

                i=i+1    
                for k in range(13, 63):
                    for j in range(1,7):
                        ws2.cell(i,1).value =  5
                        ws2.cell(i,3).value =  mack
                        ws2.cell(i,j+3).value =  ws.cell(k,j).value
                    i+=1
         wb2.close()

filename_output = "vietstock" + '_proccess.xlsx'    
wb1.save(str(filename_output))
    