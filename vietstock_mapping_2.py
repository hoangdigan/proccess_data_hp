import openpyxl as xl
import os
import glob
from openpyxl import Workbook
from openpyxl import load_workbook

filename1 ="vietstock_proccess.xlsx"
wb1 = xl.load_workbook(filename1)
ws1 = wb1['BCTC']

filename ="mapping.xlsx"
wb2 = xl.load_workbook(filename)
ws2 = wb2['MAPPING']

for i in range(2,3084):
    for j in range(1,231):
        if ws1.cell(i,1).value == ws2.cell(j,1).value and  ws1.cell(i,4).value == ws2.cell(j,4).value:
            ws1.cell(i,2).value = ws2.cell(j,2).value
            break

filename = filename1[:-5]
filename = filename +"_mapping" + ".xlsx"
wb1.save(str(filename))